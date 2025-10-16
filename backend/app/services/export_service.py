import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logger = logging.getLogger(__name__)

class ExportService:
    @staticmethod
    def create_result_excel(df: pd.DataFrame, highlight_rows: list, 
                          tu_summary: dict, tv_summary: dict, iv_summary: dict) -> BytesIO:
        """Создание Excel файла с результатами (миграция из Streamlit)"""
        try:
            logger.info("Создание Excel файла с результатами")
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Результат"

            # Запись данных
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = value

            # Подсветка строк (миграция логики из Streamlit)
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for row_idx in highlight_rows:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx + 2, column=col).fill = yellow_fill

            # Автоподбор ширины колонок
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            wb.save(output)
            output.seek(0)
            logger.info("Excel файл с результатами успешно создан")
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Ошибка создания Excel: {e}")
            raise